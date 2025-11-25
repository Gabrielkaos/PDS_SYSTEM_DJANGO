from django.shortcuts import render, redirect, HttpResponse
from .models import *
from .forms import PersonalInformationForm, FamilyBackgroundForm, EducationalBackgroundForm, OtherInformationForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.shortcuts import redirect, get_object_or_404
import openpyxl
from openpyxl.utils import get_column_letter
from .forms import ImportForm
from django.db.models import Q
from django.db.models import Count
from django.utils import timezone
from datetime import timedelta

@login_required
def dashboard(request):
    user = request.user
    
    
    total_forms = CompleteForm.objects.filter(user=user).count()
    
    
    today = timezone.now()
    first_day_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    forms_this_month = CompleteForm.objects.filter(
        user=user, 
        created_at__gte=first_day_of_month
    ).count()
    
   
    recent_forms = CompleteForm.objects.filter(user=user).order_by('-created_at')[:5]
    
    
    last_updated = CompleteForm.objects.filter(user=user).order_by('-updated_at').first()
    
    context = {
        'total_forms': total_forms,
        'forms_this_month': forms_this_month,
        'recent_forms': recent_forms,
        'last_updated': last_updated,
        "form":ImportForm(),
    }
    
    return render(request, 'pds_app/dashboard.html', context)

@login_required
def import_form(request):
    if request.method == 'POST':
        forms1 = CompleteForm.objects.filter(user=request.user)
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']

            try:
                import datetime
                from decimal import Decimal, InvalidOperation
                import openpyxl

                wb = openpyxl.load_workbook(excel_file, data_only=True)
                ws = wb.active

                print("Workbook loaded successfully")

                
                def safe_str(val):
                    return "" if val is None else str(val).strip()

                def safe_decimal(val):
                    try:
                        if val in (None, ''):
                            return None
                        return Decimal(val)
                    except InvalidOperation:
                        return None

                def safe_int(val):
                    try:
                        if val in (None, ''):
                            return None
                        return int(val)
                    except (ValueError, TypeError):
                        return None

                def safe_date(val):
                    if isinstance(val, datetime.date):
                        return val
                    if not val:
                        return None
                    try:
                        return datetime.datetime.strptime(str(val), "%Y-%m-%d").date()
                    except Exception:
                        return None

                
                def parse_vert_section(sheet, start_row=1, max_row=80):
                    data = {}
                    for row in sheet.iter_rows(min_row=start_row, max_row=max_row, max_col=2, values_only=True):
                        key, val = row
                        if key:
                            data[str(key).strip()] = val
                    return data

                data = parse_vert_section(ws)
                print("Parsed vertical section")

                form_name = safe_str(data.get("Form Name"))

                
                personal = PersonalInformation.objects.create(
                    user=request.user,
                    surname=safe_str(data.get("Surname")),
                    firstname=safe_str(data.get("First Name")),
                    middlename=safe_str(data.get("Middle Name")),
                    date_of_birth=safe_date(data.get("Date of Birth")),
                    place_of_birth=safe_str(data.get("Place of Birth")),
                    sex=safe_str(data.get("Sex")),
                    civil_status=safe_str(data.get("Civil Status")),
                    height=safe_decimal(data.get("Height (cm)")) or 0,
                    weight=safe_decimal(data.get("Weight (kg)")) or 0,
                    blood_type=safe_str(data.get("Blood Type")),
                    residential_address=safe_str(data.get("Residential Address")),
                    residential_zip_code=safe_str(data.get("Residential Zip Code")),
                    permanent_address=safe_str(data.get("Permanent Address")),
                    permanent_zip_code=safe_str(data.get("Permanent Zip Code")),
                    telephone_no=safe_str(data.get("Telephone")),
                    mobile_no=safe_str(data.get("Mobile")),
                    email_address=safe_str(data.get("Email")),
                    gsis=safe_str(data.get("GSIS")),
                    pag_ibig=safe_str(data.get("Pag-IBIG")),
                    philhealth=safe_str(data.get("PhilHealth")),
                    sss_no=safe_str(data.get("SSS")),
                    tin_no=safe_str(data.get("TIN")),
                    agent_employee_number=safe_str(data.get("Agent Employee Number")),
                    citizenship=safe_str(data.get("Citizenship")),
                )

                
                family = FamilyBackground.objects.create(
                    user=request.user,
                    spouse_surname=safe_str(data.get("Spouse Surname")),
                    spouse_firstname=safe_str(data.get("Spouse First Name")),
                    spouse_middlename=safe_str(data.get("Spouse Middle Name")),
                    spouse_name_extension=safe_str(data.get("Spouse Ext.")),
                    spouse_occupation=safe_str(data.get("Spouse Occupation")),
                    spouse_employer_business_name=safe_str(data.get("Spouse Employer")),
                    spouse_business_address=safe_str(data.get("Spouse Address")),
                    spouse_telephone_no=safe_str(data.get("Spouse Tel.")),
                    children=safe_str(data.get("Children")),
                    father_surname=safe_str(data.get("Father Surname")),
                    father_firstname=safe_str(data.get("Father First Name")),
                    father_middlename=safe_str(data.get("Father Middle Name")),
                    father_name_extension=safe_str(data.get("Father Ext.")),
                    mother_maiden_lastname=safe_str(data.get("Mother Maiden Last Name")),
                    mother_firstname=safe_str(data.get("Mother First Name")),
                    mother_middlename=safe_str(data.get("Mother Middle Name")),
                )

                
                other = OtherInformation.objects.create(
                    user=request.user,
                    with_third_degree=safe_str(data.get("With Third Degree")) or "N",
                    with_fourth_degree=safe_str(data.get("With Fourth Degree")) or "N",
                    fourth_degree_details=safe_str(data.get("Fourth Degree Details")),
                    offense=safe_str(data.get("Offense")) or "N",
                    offense_details=safe_str(data.get("Offense Details")),
                    criminal=safe_str(data.get("Criminal")) or "N",
                    criminal_details=safe_str(data.get("Criminal Details")),
                    criminal_date=safe_date(data.get("Criminal Date")),
                    convicted=safe_str(data.get("Convicted")) or "N",
                    convicted_details=safe_str(data.get("Convicted Details")),
                    sep_service=safe_str(data.get("Separated from Service")) or "N",
                    sep_service_details=safe_str(data.get("Service Separation Details")),
                    candidate=safe_str(data.get("Candidate")) or "N",
                    candidate_details=safe_str(data.get("Candidate Details")),
                    resign_candid=safe_str(data.get("Resigned for Campaign")) or "N",
                    resign_candid_details=safe_str(data.get("Campaign Resignation Details")),
                    immigrant_status=safe_str(data.get("Immigration Status")) or "N",
                    immigrant_details=safe_str(data.get("Immigration Details")),
                    indigenous_group_member=safe_str(data.get("Indigenous Member")) or "N",
                    disability_status=safe_str(data.get("Disability Status")) or "N",
                    solo_parent_status=safe_str(data.get("Solo Parent")) or "N",
                    references=safe_str(data.get("References")),
                    government_id=safe_str(data.get("Government ID")),
                    government_id_number=safe_str(data.get("Government ID Number")),
                    id_issue_date=safe_date(data.get("ID Issue Date")),
                    id_issue_place=safe_str(data.get("ID Issue Place")),
                )

                
                def find_section_start(ws, title):
                    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                        if row and row[0] and title.strip().upper() in str(row[0]).upper():
                            return i
                    return None

                def parse_table(ws, start_row, num_cols):
                    data_list = []
                    for row in ws.iter_rows(min_row=start_row + 1, values_only=True):
                        if not any(row):
                            break
                        data_list.append(list(row[:num_cols]))
                    return data_list

                
                idx = find_section_start(ws, "EDUCATIONAL BACKGROUND")
                educational = None
                if idx:
                    edu_data = parse_table(ws, idx + 1, 8)
                    edu_dict = {}
                    for row in edu_data:
                        if not row[0]:
                            continue
                        level = str(row[0]).strip().lower()
                        edu_dict[level] = row[1:]

                    educational = EducationalBackground.objects.create(
                        user=request.user,
                        elementary_name=safe_str(edu_dict.get("elementary", [""])[0]),
                        elementary_degree_course=safe_str(edu_dict.get("elementary", ["", ""])[1]),
                        elementary_period_from=safe_str(edu_dict.get("elementary", ["", "", ""])[2]),
                        elementary_period_to=safe_str(edu_dict.get("elementary", ["", "", "", ""])[3]),
                        elementary_highest_level_units=safe_str(edu_dict.get("elementary", ["", "", "", "", ""])[4]),
                        elementary_year_graduated=safe_str(edu_dict.get("elementary", ["", "", "", "", "", ""])[5]),
                        elementary_honors=safe_str(edu_dict.get("elementary", ["", "", "", "", "", "", ""])[6]),
                        secondary_name=safe_str(edu_dict.get("secondary", [""])[0]),
                        secondary_degree_course=safe_str(edu_dict.get("secondary", ["", ""])[1]),
                        secondary_period_from=safe_str(edu_dict.get("secondary", ["", "", ""])[2]),
                        secondary_period_to=safe_str(edu_dict.get("secondary", ["", "", "", ""])[3]),
                        secondary_highest_level_units=safe_str(edu_dict.get("secondary", ["", "", "", "", ""])[4]),
                        secondary_year_graduated=safe_str(edu_dict.get("secondary", ["", "", "", "", "", ""])[5]),
                        secondary_honors=safe_str(edu_dict.get("secondary", ["", "", "", "", "", "", ""])[6]),
                        vocational_name=safe_str(edu_dict.get("vocational", [""])[0]),
                        vocational_degree_course=safe_str(edu_dict.get("vocational", ["", ""])[1]),
                        vocational_period_from=safe_str(edu_dict.get("vocational", ["", "", ""])[2]),
                        vocational_period_to=safe_str(edu_dict.get("vocational", ["", "", "", ""])[3]),
                        vocational_highest_level_units=safe_str(edu_dict.get("vocational", ["", "", "", "", ""])[4]),
                        vocational_year_graduated=safe_str(edu_dict.get("vocational", ["", "", "", "", "", ""])[5]),
                        vocational_honors=safe_str(edu_dict.get("vocational", ["", "", "", "", "", "", ""])[6]),
                        college_name=safe_str(edu_dict.get("college", [""])[0]),
                        college_degree_course=safe_str(edu_dict.get("college", ["", ""])[1]),
                        college_period_from=safe_str(edu_dict.get("college", ["", "", ""])[2]),
                        college_period_to=safe_str(edu_dict.get("college", ["", "", "", ""])[3]),
                        college_highest_level_units=safe_str(edu_dict.get("college", ["", "", "", "", ""])[4]),
                        college_year_graduated=safe_str(edu_dict.get("college", ["", "", "", "", "", ""])[5]),
                        college_honors=safe_str(edu_dict.get("college", ["", "", "", "", "", "", ""])[6]),
                        graduate_name=safe_str(edu_dict.get("graduate studies", [""])[0]),
                        graduate_degree_course=safe_str(edu_dict.get("graduate studies", ["", ""])[1]),
                        graduate_period_from=safe_str(edu_dict.get("graduate studies", ["", "", ""])[2]),
                        graduate_period_to=safe_str(edu_dict.get("graduate studies", ["", "", "", ""])[3]),
                        graduate_highest_level_units=safe_str(edu_dict.get("graduate studies", ["", "", "", "", ""])[4]),
                        graduate_year_graduated=safe_str(edu_dict.get("graduate studies", ["", "", "", "", "", ""])[5]),
                        graduate_honors=safe_str(edu_dict.get("graduate studies", ["", "", "", "", "", "", ""])[6]),
                    )
                else:
                    messages.warning(request, "Educational background section not found.")
                    educational = EducationalBackground.objects.create(user=request.user)

                
                civil_objs = []
                idx = find_section_start(ws, "CIVIL SERVICE ELIGIBILITY")
                if idx:
                    for row in parse_table(ws, idx + 1, 6):
                        civil_objs.append(CivilServiceEligibility.objects.create(
                            user=request.user,
                            career_service=safe_str(row[0]),
                            rating=safe_decimal(row[1]),
                            exam_date=safe_date(row[2]),
                            exam_place=safe_str(row[3]),
                            license_number=safe_str(row[4]),
                            license_validity=safe_date(row[5]),
                        ))

                
                work_objs = []
                idx = find_section_start(ws, "WORK EXPERIENCE")
                if idx:
                    for row in parse_table(ws, idx + 1, 8):
                        work_objs.append(WorkExperience.objects.create(
                            user=request.user,
                            from_date=safe_date(row[0]),
                            to_date=safe_date(row[1]),
                            position_title=safe_str(row[2]),
                            department=safe_str(row[3]),
                            monthly_salary=safe_decimal(row[4]),
                            salary_grade=safe_str(row[5]),
                            status_of_appointment=safe_str(row[6]),
                            govt_service=safe_str(row[7]) or "N",
                        ))

                
                vol_objs = []
                idx = find_section_start(ws, "VOLUNTARY WORK")
                if idx:
                    for row in parse_table(ws, idx + 1, 5):
                        vol_objs.append(VoluntaryWork.objects.create(
                            user=request.user,
                            organization_name=safe_str(row[0]),
                            from_date=safe_date(row[1]),
                            to_date=safe_date(row[2]),
                            number_of_hours=safe_int(row[3]),
                            nature_of_work=safe_str(row[4]),
                        ))

                
                ld_objs = []
                idx = find_section_start(ws, "LEARNING AND DEVELOPMENT")
                if idx:
                    for row in parse_table(ws, idx + 1, 6):
                        ld_objs.append(LearningDevelopment.objects.create(
                            user=request.user,
                            title=safe_str(row[0]),
                            from_date=safe_date(row[1]),
                            to_date=safe_date(row[2]),
                            number_of_hours=safe_int(row[3]),
                            type_of_ld=safe_str(row[4]),
                            conducted_by=safe_str(row[5]),
                        ))

                
                complete_form = CompleteForm.objects.create(
                    user=request.user,
                    name=form_name,
                    personal_information=personal,
                    family_background=family,
                    other_information=other,
                    educational_background=educational,
                )
                complete_form.civil_service.set(civil_objs)
                complete_form.work_experience.set(work_objs)
                complete_form.voluntary_work.set(vol_objs)
                complete_form.learning_development.set(ld_objs)

                messages.success(request, "Form imported successfully!")
                return redirect('edit_form',form_id=complete_form.id)

            except Exception as e:
                import traceback
                traceback.print_exc()
                messages.error(request, f"Failed to import Excel: {e}")
                return redirect("create_form")

    else:
        forms1 = CompleteForm.objects.filter(user=request.user)
        form = ImportForm()

    return render(request, 'pds_app/forms.html', {'form': form, "forms": forms1})


@login_required
def export_form(request, form_id):
    form = get_object_or_404(CompleteForm, id=form_id, user=request.user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personal Data Sheet"

    row = 1

    
    def write(label, value):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=str(value) if value is not None else "")
        row += 1

    #form name first
    write("Form Name",form.name)

    row+=1
    write("personalinformation","")

    p = form.personal_information
    write("Surname", p.surname)
    write("First Name", p.firstname)
    write("Middle Name", p.middlename)
    write("Date of Birth", p.date_of_birth)
    write("Place of Birth", p.place_of_birth)
    write("Sex", p.sex)
    write("Civil Status", p.civil_status)
    write("Height (cm)", p.height)
    write("Weight (kg)", p.weight)
    write("Blood Type", p.blood_type)
    write("Residential Address", p.residential_address)
    write("Residential Zip Code", p.residential_zip_code)
    write("Permanent Address", p.permanent_address)
    write("Permanent Zip Code", p.permanent_zip_code)
    write("Telephone", p.telephone_no)
    write("Mobile", p.mobile_no)
    write("Email", p.email_address)
    write("GSIS", p.gsis)
    write("Pag-IBIG", p.pag_ibig)
    write("PhilHealth", p.philhealth)
    write("SSS", p.sss_no)
    write("TIN", p.tin_no)
    write("Agent Employee Number", p.agent_employee_number)
    write("Citizenship", p.citizenship)

    row += 1
    write("familybackground", "")

    f = form.family_background
    write("Spouse Surname", f.spouse_surname)
    write("Spouse First Name", f.spouse_firstname)
    write("Spouse Middle Name", f.spouse_middlename)
    write("Spouse Ext.", f.spouse_name_extension)
    write("Spouse Occupation", f.spouse_occupation)
    write("Spouse Employer", f.spouse_employer_business_name)
    write("Spouse Address", f.spouse_business_address)
    write("Spouse Tel.", f.spouse_telephone_no)
    write("Children", f.children)
    write("Father Surname", f.father_surname)
    write("Father First Name", f.father_firstname)
    write("Father Middle Name", f.father_middlename)
    write("Father Ext.", f.father_name_extension)
    write("Mother Maiden Last Name", f.mother_maiden_lastname)
    write("Mother First Name", f.mother_firstname)
    write("Mother Middle Name", f.mother_middlename)

    row += 1
    write("otherinformation", "")

    o = form.other_information
    write("With Third Degree", o.with_third_degree)
    write("With Fourth Degree", o.with_fourth_degree)
    write("Fourth Degree Details", o.fourth_degree_details)
    write("Offense", o.offense)
    write("Offense Details", o.offense_details)
    write("Criminal", o.criminal)
    write("Criminal Details", o.criminal_details)
    write("Criminal Date", o.criminal_date)
    write("Convicted", o.convicted)
    write("Convicted Details", o.convicted_details)
    write("Separated from Service", o.sep_service)
    write("Service Separation Details", o.sep_service_details)
    write("Candidate", o.candidate)
    write("Candidate Details", o.candidate_details)
    write("Resigned for Campaign", o.resign_candid)
    write("Campaign Resignation Details", o.resign_candid_details)
    write("Immigration Status", o.immigrant_status)
    write("Immigration Details", o.immigrant_details)
    write("Indigenous Member", o.indigenous_group_member)
    write("Disability Status", o.disability_status)
    write("Solo Parent", o.solo_parent_status)
    write("References", o.references)
    write("Government ID", o.government_id)
    write("Government ID Number", o.government_id_number)
    write("ID Issue Date", o.id_issue_date)
    write("ID Issue Place", o.id_issue_place)

    row += 2

    ws.cell(row=row, column=1, value="EDUCATIONAL BACKGROUND"); row += 1
    e = form.educational_background

    ws.append(["Level", "Name of School", "Degree/Course", "From", "To", "Highest Level/Units", "Year Graduated", "Honors"])
    row += 1
    ws.append(["Elementary", e.elementary_name, e.elementary_degree_course, e.elementary_period_from, e.elementary_period_to, e.elementary_highest_level_units, e.elementary_year_graduated, e.elementary_honors])
    ws.append(["Secondary", e.secondary_name, e.secondary_degree_course, e.secondary_period_from, e.secondary_period_to, e.secondary_highest_level_units, e.secondary_year_graduated, e.secondary_honors])
    ws.append(["Vocational", e.vocational_name, e.vocational_degree_course, e.vocational_period_from, e.vocational_period_to, e.vocational_highest_level_units, e.vocational_year_graduated, e.vocational_honors])
    ws.append(["College", e.college_name, e.college_degree_course, e.college_period_from, e.college_period_to, e.college_highest_level_units, e.college_year_graduated, e.college_honors])
    ws.append(["Graduate Studies", e.graduate_name, e.graduate_degree_course, e.graduate_period_from, e.graduate_period_to, e.graduate_highest_level_units, e.graduate_year_graduated, e.graduate_honors])

    row = ws.max_row + 2

    ws.cell(row=row, column=1, value="CIVIL SERVICE ELIGIBILITY"); row += 1
    ws.append(["Career Service", "Rating", "Exam Date", "Exam Place", "License #", "License Validity"])
    for item in form.civil_service.all():
        ws.append([item.career_service, item.rating, item.exam_date, item.exam_place, item.license_number, item.license_validity])
    row = ws.max_row + 2

    ws.cell(row=row, column=1, value="WORK EXPERIENCE"); row += 1
    ws.append(["From", "To", "Position Title", "Department/Agency", "Monthly Salary", "Salary Grade & STEP", "Status of Appointment", "Govt Service"])
    for w in form.work_experience.all():
        ws.append([w.from_date, w.to_date, w.position_title, w.department, w.monthly_salary, w.salary_grade, w.status_of_appointment, w.govt_service])
    row = ws.max_row + 2

    ws.cell(row=row, column=1, value="VOLUNTARY WORK"); row += 1
    ws.append(["Organization", "From", "To", "Hours", "Nature"])
    for v in form.voluntary_work.all():
        ws.append([v.organization_name, v.from_date, v.to_date, v.number_of_hours, v.nature_of_work])
    row = ws.max_row + 2

    ws.cell(row=row, column=1, value="LEARNING AND DEVELOPMENT"); row += 1
    ws.append(["Title", "From", "To", "Hours", "Type", "Conducted By"])
    for l in form.learning_development.all():
        ws.append([l.title, l.from_date, l.to_date, l.number_of_hours, l.type_of_ld, l.conducted_by])
    row = ws.max_row + 2

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'PDS_{form.name}_{form.id}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@login_required
def home(request):
    forms_list = CompleteForm.objects.filter(user=request.user).select_related(
        'personal_information'
    )
    
    search_query = request.GET.get('search', '').strip()
    if search_query:
        forms_list = forms_list.filter(
            Q(name__icontains=search_query) |
            Q(personal_information__surname__icontains=search_query) |
            Q(personal_information__firstname__icontains=search_query) |
            Q(personal_information__email_address__icontains=search_query) |
            Q(personal_information__mobile_no__icontains=search_query)
        ).distinct()
    
    sort_by = request.GET.get('sort', '-created_at')
    if sort_by in ['name', '-name', 'created_at', '-created_at']:
        forms_list = forms_list.order_by(sort_by)
    
    form = ImportForm()
    
    context = {
        'forms': forms_list,
        'form': form,
        'search_query': search_query,
        'total_forms' : forms_list.__len__()
    }
    
    return render(request, "pds_app/forms.html", context)

@login_required
def all_forms(request, form_id):
    """View specific form - ensure user can only see their own forms"""
    form_instance = get_object_or_404(CompleteForm, id=form_id, user=request.user)
    user_forms = CompleteForm.objects.filter(user=request.user)
    
    return render(request, "pds_app/all_forms.html", {
        "this_form": form_instance, 
        "forms": user_forms,
        "form":ImportForm()
    })

@login_required
def forms(request):
    user_forms = CompleteForm.objects.filter(user=request.user)
    form = ImportForm()
    return render(request, "pds_app/forms.html", {"forms": user_forms,"form":form})

@login_required
def delete_form(request, form_id):
    form = get_object_or_404(CompleteForm, id=form_id, user=request.user)
    
    if request.method == "POST":
        form.delete()
        return redirect('home')
    
    return redirect('home')

@login_required
def logout_view(request):
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect('login')

@login_required
def edit_form(request, form_id):
    form_instance = get_object_or_404(CompleteForm, id=form_id, user=request.user)
    
    if request.method == 'POST':
        print(request.POST)
        print("POST")
        personal_form = PersonalInformationForm(request.POST, instance=form_instance.personal_information)
        family_form = FamilyBackgroundForm(request.POST, instance=form_instance.family_background)
        education_form = EducationalBackgroundForm(request.POST, instance=form_instance.educational_background)
        other_form = OtherInformationForm(request.POST, instance=form_instance.other_information)

        print("Personal form errors:", personal_form.errors)
        print("Family form errors:", family_form.errors)
        print("Education form errors:", education_form.errors)
        print("Other form errors:", other_form.errors)

        print([personal_form.is_valid(), family_form.is_valid(), education_form.is_valid(), other_form.is_valid()])
        
        if all([personal_form.is_valid(), family_form.is_valid(), other_form.is_valid(),education_form.is_valid()]):

            personal_form.save()
            family_form.save()
            education_form.save()
            other_form.save()
            print("Saved")

            form_instance.civil_service.all().delete()

            career_services = request.POST.getlist('career_service[]')
            ratings = request.POST.getlist('rating[]')
            exam_dates = request.POST.getlist('exam_date[]')
            exam_places = request.POST.getlist('exam_place[]')
            license_numbers = request.POST.getlist('license_number[]')
            license_validities = request.POST.getlist('license_validity[]')

            print("career_services:", request.POST.getlist('career_service[]'))
            print("ratings:", request.POST.getlist('rating[]'))
            print("exam_dates:", request.POST.getlist('exam_date[]'))
            print("exam_places:", request.POST.getlist('exam_place[]'))
            print("license_number:", request.POST.getlist('license_number[]'))
            print("license_validities:", request.POST.getlist('license_validity[]'))

            civil_services = []
            for i in range(len(career_services)):
                if career_services[i].strip():
                    civil_service = CivilServiceEligibility.objects.create(
                        user=request.user,
                        career_service=career_services[i],
                        rating=ratings[i] if i < len(ratings) else None,
                        exam_date=exam_dates[i] if i < len(exam_dates) else None,
                        exam_place=exam_places[i] if i < len(exam_places) else None,
                        license_number=license_numbers[i] if i < len(license_numbers) else None,
                        license_validity=license_validities[i] if i < len(license_validities) else None,
                    )
                    civil_services.append(civil_service)
            form_instance.civil_service.set(civil_services)


            form_instance.work_experience.all().delete()
            work_from_dates = request.POST.getlist('work_from_date[]')
            work_experiences = []
            for i in range(len(work_from_dates)):
                if work_from_dates[i]:
                    work_exp = WorkExperience.objects.create(
                        user=request.user,
                        from_date=work_from_dates[i],
                        to_date=request.POST.getlist('work_to_date[]')[i] if i < len(request.POST.getlist('work_to_date[]')) and request.POST.getlist('work_to_date[]')[i] else None,
                        position_title=request.POST.getlist('work_position_title[]')[i] if i < len(request.POST.getlist('work_position_title[]')) else '',
                        department=request.POST.getlist('work_department[]')[i] if i < len(request.POST.getlist('work_department[]')) else '',
                        monthly_salary=request.POST.getlist('work_monthly_salary[]')[i] if i < len(request.POST.getlist('work_monthly_salary[]')) and request.POST.getlist('work_monthly_salary[]')[i] else None,
                        salary_grade=request.POST.getlist('work_salary_grade[]')[i] if i < len(request.POST.getlist('work_salary_grade[]')) else None,
                        status_of_appointment=request.POST.getlist('work_status[]')[i] if i < len(request.POST.getlist('work_status[]')) else '',
                        govt_service=request.POST.getlist('work_govt_service[]')[i] if i < len(request.POST.getlist('work_govt_service[]')) else 'N',
                    )
                    work_experiences.append(work_exp)
            form_instance.work_experience.set(work_experiences)

            form_instance.voluntary_work.all().delete()
            voluntary_orgs = request.POST.getlist('voluntary_org[]')
            voluntary_works = []
            for i in range(len(voluntary_orgs)):
                if voluntary_orgs[i].strip():
                    voluntary = VoluntaryWork.objects.create(
                        user=request.user,
                        organization_name=voluntary_orgs[i],
                        from_date=request.POST.getlist('voluntary_from[]')[i] if i < len(request.POST.getlist('voluntary_from[]')) and request.POST.getlist('voluntary_from[]')[i] else None,
                        to_date=request.POST.getlist('voluntary_to[]')[i] if i < len(request.POST.getlist('voluntary_to[]')) and request.POST.getlist('voluntary_to[]')[i] else None,
                        number_of_hours=request.POST.getlist('voluntary_hours[]')[i] if i < len(request.POST.getlist('voluntary_hours[]')) and request.POST.getlist('voluntary_hours[]')[i] else 0,
                        nature_of_work=request.POST.getlist('voluntary_nature[]')[i] if i < len(request.POST.getlist('voluntary_nature[]')) else '',
                    )
                    voluntary_works.append(voluntary)
            form_instance.voluntary_work.set(voluntary_works)

            form_instance.learning_development.all().delete()
            learning_titles = request.POST.getlist('learning_title[]')

            learning_developments = []
            for i in range(len(learning_titles)):
                if learning_titles[i].strip():
                    learning = LearningDevelopment.objects.create(
                        user=request.user,
                        title=learning_titles[i],
                        from_date=request.POST.getlist('learning_from[]')[i] if i < len(request.POST.getlist('learning_from[]')) and request.POST.getlist('learning_from[]')[i] else None,
                        to_date=request.POST.getlist('learning_to[]')[i] if i < len(request.POST.getlist('learning_to[]')) and request.POST.getlist('learning_to[]')[i] else None,
                        number_of_hours=request.POST.getlist('learning_hours[]')[i] if i < len(request.POST.getlist('learning_hours[]')) and request.POST.getlist('learning_hours[]')[i] else 0,
                        type_of_ld=request.POST.getlist('learning_type[]')[i] if i < len(request.POST.getlist('learning_type[]')) else '',
                        conducted_by=request.POST.getlist('learning_conducted[]')[i] if i < len(request.POST.getlist('learning_conducted[]')) else '',
                    )
                    learning_developments.append(learning)
            form_instance.learning_development.set(learning_developments)

                        
            if 'form_name' in request.POST:
                form_instance.name = request.POST["form_name"]
                form_instance.save()
                print("saved form name")


            messages.success(request, "Form edited successfully!")
            return redirect('all_forms', form_id=form_id)


        else:
            for form in [personal_form, family_form, education_form, other_form]:
                if form.errors:
                    for field, errors in form.errors.items():
                        for error in errors:
                            messages.error(request, f"{form[field].label}: {error}")
    else:
        personal_form = PersonalInformationForm(instance=form_instance.personal_information)
        family_form = FamilyBackgroundForm(instance=form_instance.family_background)
        education_form = EducationalBackgroundForm(instance=form_instance.educational_background)
        other_form = OtherInformationForm(instance=form_instance.other_information)
    
    print(form_instance.civil_service.all)
    context = {
        'form_instance': form_instance,
        'personal_form': personal_form,
        'family_form': family_form,
        'education_form': education_form,
        'other_form': other_form,
        'creating': False,
        'form':ImportForm(),
        "forms" : CompleteForm.objects.filter(user=request.user)
    }
    
    return render(request, 'pds_app/edit_form.html', context)

@login_required
def create_form(request):
    if request.method == 'POST':
        form_name = request.POST.get("form_name")
        print(form_name)
        personal_form = PersonalInformationForm(request.POST)
        family_form = FamilyBackgroundForm(request.POST)
        education_form = EducationalBackgroundForm(request.POST)
        other_form = OtherInformationForm(request.POST)

        print("Personal form errors:", personal_form.errors)
        print("Family form errors:", family_form.errors)
        print("Education form errors:", education_form.errors)
        print("Other form errors:", other_form.errors)

        print([personal_form.is_valid(), family_form.is_valid(), education_form.is_valid(), other_form.is_valid()])
        
        if all([personal_form.is_valid(), family_form.is_valid(), other_form.is_valid(),education_form.is_valid()]):

            personal_inst = personal_form.save(commit=False)
            family_inst = family_form.save(commit=False)
            education_inst = education_form.save(commit=False)
            other_inst = other_form.save(commit=False)

            for inst in [personal_inst, family_inst, education_inst, other_inst]:
                inst.user = request.user
                inst.save()
            print("Saved")

        
            career_services = request.POST.getlist('career_service[]')
            ratings = request.POST.getlist('rating[]')
            exam_dates = request.POST.getlist('exam_date[]')
            exam_places = request.POST.getlist('exam_place[]')
            license_numbers = request.POST.getlist('license_number[]')
            license_validities = request.POST.getlist('license_validity[]')
            civil_services = []
            for i in range(len(career_services)):
                if career_services[i].strip():
                    civil_service = CivilServiceEligibility.objects.create(
                        user=request.user,
                        career_service=career_services[i],
                        rating=ratings[i] if i < len(ratings) else None,
                        exam_date=exam_dates[i] if i < len(exam_dates) else None,
                        exam_place=exam_places[i] if i < len(exam_places) else None,
                        license_number=license_numbers[i] if i < len(license_numbers) else None,
                        license_validity=license_validities[i] if i < len(license_validities) else None,
                    )
                    civil_services.append(civil_service)
            

            work_from_dates = request.POST.getlist('work_from_date[]')
            work_experiences = []
            for i in range(len(work_from_dates)):
                if work_from_dates[i]:
                    work_exp = WorkExperience.objects.create(
                        user=request.user,
                        from_date=work_from_dates[i],
                        to_date=request.POST.getlist('work_to_date[]')[i] if i < len(request.POST.getlist('work_to_date[]')) and request.POST.getlist('work_to_date[]')[i] else None,
                        position_title=request.POST.getlist('work_position_title[]')[i] if i < len(request.POST.getlist('work_position_title[]')) else '',
                        department=request.POST.getlist('work_department[]')[i] if i < len(request.POST.getlist('work_department[]')) else '',
                        monthly_salary=request.POST.getlist('work_monthly_salary[]')[i] if i < len(request.POST.getlist('work_monthly_salary[]')) and request.POST.getlist('work_monthly_salary[]')[i] else None,
                        salary_grade=request.POST.getlist('work_salary_grade[]')[i] if i < len(request.POST.getlist('work_salary_grade[]')) else None,
                        status_of_appointment=request.POST.getlist('work_status[]')[i] if i < len(request.POST.getlist('work_status[]')) else '',
                        govt_service=request.POST.getlist('work_govt_service[]')[i] if i < len(request.POST.getlist('work_govt_service[]')) else 'N',
                    )
                    work_experiences.append(work_exp)

            voluntary_orgs = request.POST.getlist('voluntary_org[]')
            voluntary_works = []
            for i in range(len(voluntary_orgs)):
                if voluntary_orgs[i].strip():
                    voluntary = VoluntaryWork.objects.create(
                        user=request.user,
                        organization_name=voluntary_orgs[i],
                        from_date=request.POST.getlist('voluntary_from[]')[i] if i < len(request.POST.getlist('voluntary_from[]')) and request.POST.getlist('voluntary_from[]')[i] else None,
                        to_date=request.POST.getlist('voluntary_to[]')[i] if i < len(request.POST.getlist('voluntary_to[]')) and request.POST.getlist('voluntary_to[]')[i] else None,
                        number_of_hours=request.POST.getlist('voluntary_hours[]')[i] if i < len(request.POST.getlist('voluntary_hours[]')) and request.POST.getlist('voluntary_hours[]')[i] else 0,
                        nature_of_work=request.POST.getlist('voluntary_nature[]')[i] if i < len(request.POST.getlist('voluntary_nature[]')) else '',
                    )
                    voluntary_works.append(voluntary)

            learning_titles = request.POST.getlist('learning_title[]')
            learning_developments = []
            for i in range(len(learning_titles)):
                if learning_titles[i].strip():
                    learning = LearningDevelopment.objects.create(
                        user=request.user,
                        title=learning_titles[i],
                        from_date=request.POST.getlist('learning_from[]')[i] if i < len(request.POST.getlist('learning_from[]')) and request.POST.getlist('learning_from[]')[i] else None,
                        to_date=request.POST.getlist('learning_to[]')[i] if i < len(request.POST.getlist('learning_to[]')) and request.POST.getlist('learning_to[]')[i] else None,
                        number_of_hours=request.POST.getlist('learning_hours[]')[i] if i < len(request.POST.getlist('learning_hours[]')) and request.POST.getlist('learning_hours[]')[i] else 0,
                        type_of_ld=request.POST.getlist('learning_type[]')[i] if i < len(request.POST.getlist('learning_type[]')) else '',
                        conducted_by=request.POST.getlist('learning_conducted[]')[i] if i < len(request.POST.getlist('learning_conducted[]')) else '',
                    )
                    learning_developments.append(learning)

            complete_form = CompleteForm.objects.create(
                user=request.user,
                name=form_name,
                personal_information=personal_inst,
                family_background=family_inst,
                educational_background=education_inst,
                other_information=other_inst
            )
            complete_form.civil_service.set(civil_services)
            complete_form.work_experience.set(work_experiences)
            complete_form.voluntary_work.set(voluntary_works)
            complete_form.learning_development.set(learning_developments)

            messages.success(request, "Form created successfully!")
            return redirect('edit_form',form_id=complete_form.id)
        else:
            for form in [personal_form, family_form, education_form, other_form]:
                if form.errors:
                    for field, errors in form.errors.items():
                        for error in errors:
                            messages.error(request, f"{form[field].label}: {error}")

    else:
        personal_form = PersonalInformationForm()
        family_form = FamilyBackgroundForm()
        education_form = EducationalBackgroundForm()
        other_form = OtherInformationForm()

    context = {
        'personal_form': personal_form,
        'family_form': family_form,
        'education_form': education_form,
        'other_form': other_form,
        'creating': True,
        "forms": CompleteForm.objects.filter(user=request.user),
        "form":ImportForm()
    }

    return render(request, 'pds_app/edit_form.html', context)
